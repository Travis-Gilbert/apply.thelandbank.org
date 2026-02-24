"""
Shared import utilities for Property records.

Used by both management commands (CLI) and admin actions (UI).
Supports two formats:
  - CSV: flexible column mapping, dedup by parcel_id
  - Excel (.xlsx): FileMaker export format, filters to GCLB-owned properties,
    infers program type from Structure Flag

Both support optional replace mode that withdraws existing available properties.
"""

import csv
import io
from decimal import Decimal, InvalidOperation

from django.utils import timezone

from .models import Application, Property

# Flexible column name mapping — accepts common variations
COLUMN_MAP = {
    # address
    "address": "address",
    "property_address": "address",
    "property address": "address",
    "street_address": "address",
    "street address": "address",
    # parcel_id
    "parcel_id": "parcel_id",
    "parcel id": "parcel_id",
    "parcelid": "parcel_id",
    "parcel": "parcel_id",
    "pid": "parcel_id",
    # program
    "program": "program_type",
    "program_type": "program_type",
    "program type": "program_type",
    # listing_price
    "listing_price": "listing_price",
    "listing price": "listing_price",
    "price": "listing_price",
    "list_price": "listing_price",
    "list price": "listing_price",
}

# Map common program name variations to ProgramType values
PROGRAM_ALIASES = {
    "featured_homes": Application.ProgramType.FEATURED_HOMES,
    "featured homes": Application.ProgramType.FEATURED_HOMES,
    "featured": Application.ProgramType.FEATURED_HOMES,
    "fh": Application.ProgramType.FEATURED_HOMES,
    "ready_for_rehab": Application.ProgramType.READY_FOR_REHAB,
    "ready for rehab": Application.ProgramType.READY_FOR_REHAB,
    "r4r": Application.ProgramType.READY_FOR_REHAB,
    "rehab": Application.ProgramType.READY_FOR_REHAB,
    "vip_spotlight": Application.ProgramType.VIP_SPOTLIGHT,
    "vip spotlight": Application.ProgramType.VIP_SPOTLIGHT,
    "vip": Application.ProgramType.VIP_SPOTLIGHT,
    "spotlight": Application.ProgramType.VIP_SPOTLIGHT,
    "vacant_lot": Application.ProgramType.VACANT_LOT,
    "vacant lot": Application.ProgramType.VACANT_LOT,
    "vacant": Application.ProgramType.VACANT_LOT,
    "lot": Application.ProgramType.VACANT_LOT,
}


def _resolve_columns(headers):
    """Map CSV headers to internal field names. Returns {field_name: csv_index}."""
    mapping = {}
    for idx, raw_header in enumerate(headers):
        normalized = raw_header.strip().lower().replace("-", "_")
        field = COLUMN_MAP.get(normalized)
        if field and field not in mapping:
            mapping[field] = idx
    return mapping


def _parse_price(value):
    """Parse a price string like '$12,500.00' or '12500' into Decimal or None."""
    if not value:
        return None
    cleaned = value.strip().replace("$", "").replace(",", "")
    if not cleaned:
        return None
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _resolve_program(value):
    """Convert a program string to a ProgramType value, or None."""
    if not value:
        return None
    normalized = value.strip().lower()
    return PROGRAM_ALIASES.get(normalized)


def import_properties_from_csv(csv_file, replace_existing=False, batch_label=""):
    """
    Import properties from a CSV file.

    Args:
        csv_file: A file-like object (text mode) or Django UploadedFile.
        replace_existing: If True, mark current 'available' properties as
                         'withdrawn' before import.
        batch_label: Optional label for this import batch.

    Returns:
        dict with keys: created, updated, skipped, errors (list of strings)
    """
    result = {"created": 0, "updated": 0, "skipped": 0, "errors": []}

    # Handle Django UploadedFile (binary) vs text file
    if hasattr(csv_file, "read"):
        raw = csv_file.read()
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8-sig")  # Handle BOM from Excel
        text = raw
    else:
        text = csv_file

    reader = csv.reader(io.StringIO(text))

    try:
        headers = next(reader)
    except StopIteration:
        result["errors"].append("CSV file is empty.")
        return result

    col_map = _resolve_columns(headers)

    # Validate required columns
    missing = []
    for required in ("address", "parcel_id", "program_type"):
        if required not in col_map:
            missing.append(required)
    if missing:
        result["errors"].append(
            f"Missing required columns: {', '.join(missing)}. "
            f"Found columns: {', '.join(h.strip() for h in headers)}"
        )
        return result

    if not batch_label:
        batch_label = f"import-{timezone.now().strftime('%Y%m%d-%H%M%S')}"

    # Optionally withdraw existing available properties
    if replace_existing:
        Property.objects.filter(status=Property.Status.AVAILABLE).update(
            status=Property.Status.WITHDRAWN
        )

    for row_num, row in enumerate(reader, start=2):
        if not any(cell.strip() for cell in row):
            continue  # Skip blank rows

        try:
            address = row[col_map["address"]].strip()
            parcel_id = row[col_map["parcel_id"]].strip()
            program_raw = row[col_map["program_type"]].strip()
        except IndexError:
            result["errors"].append(f"Row {row_num}: insufficient columns")
            continue

        if not address:
            result["errors"].append(f"Row {row_num}: missing address")
            continue
        if not parcel_id:
            result["errors"].append(f"Row {row_num}: missing parcel_id")
            continue

        program_type = _resolve_program(program_raw)
        if not program_type:
            result["errors"].append(
                f"Row {row_num}: unrecognized program '{program_raw}'"
            )
            continue

        listing_price = None
        if "listing_price" in col_map:
            listing_price = _parse_price(row[col_map["listing_price"]])

        # Upsert by parcel_id
        try:
            prop, created = Property.objects.update_or_create(
                parcel_id=parcel_id,
                defaults={
                    "address": address,
                    "program_type": program_type,
                    "listing_price": listing_price,
                    "status": Property.Status.AVAILABLE,
                    "csv_batch": batch_label,
                },
            )
            if created:
                result["created"] += 1
            else:
                result["updated"] += 1
        except Exception as e:
            result["errors"].append(f"Row {row_num}: {e}")

    return result


# ── Excel (FileMaker) import ────────────────────────────────────────


def _infer_program_type(program_raw, structure_flag):
    """
    Determine program type from explicit column or Structure Flag.

    Priority: explicit Program Type > Structure Flag inference.
    Structure=Yes → featured_homes (has a house)
    Structure=No  → vacant_lot (empty lot)

    Returns None if neither source provides a valid program type.
    """
    if program_raw:
        resolved = _resolve_program(program_raw)
        if resolved:
            return resolved
        # Explicit value was set but unrecognized — return None so caller logs error
        return None

    if structure_flag == "Yes":
        return Application.ProgramType.FEATURED_HOMES
    elif structure_flag == "No":
        return Application.ProgramType.VACANT_LOT

    # Neither program type nor structure flag available
    return None


def import_properties_from_excel(excel_file, replace_existing=False, batch_label=""):
    """
    Import GCLB-owned properties from a FileMaker Excel export.

    Expected columns (by header name):
        Street Address, City State Zip, GCLB Owned, Structure Flag,
        Minimum Bid, Program Type

    Only rows where "GCLB Owned" == "Yes" are imported.
    Program type is inferred from Structure Flag when not explicitly set.
    Deduplication on reimport uses address_normalized since no parcel_id
    is available in the FileMaker export.

    Args:
        excel_file: File path (str) or file-like object (Django UploadedFile).
        replace_existing: If True, withdraw current available properties first.
        batch_label: Optional label for this import batch.

    Returns:
        dict with keys: created, updated, skipped, errors (list of strings)
    """
    import openpyxl

    result = {"created": 0, "updated": 0, "skipped": 0, "errors": []}

    # Load workbook — handle file path or file-like object
    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True)
    except Exception as e:
        result["errors"].append(f"Could not open Excel file: {e}")
        return result

    # Find the right sheet — try "Properties" first, fall back to first sheet
    if "Properties" in wb.sheetnames:
        ws = wb["Properties"]
    else:
        ws = wb[wb.sheetnames[0]]

    # Read headers from first row
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        header_row = row
        break

    if not header_row:
        result["errors"].append("Excel file is empty or has no header row.")
        wb.close()
        return result

    # Build column index map from headers
    col_idx = {}
    for i, h in enumerate(header_row):
        if h:
            col_idx[str(h).strip()] = i

    # Validate required columns
    required = ["Street Address"]
    for col_name in required:
        if col_name not in col_idx:
            result["errors"].append(
                f"Missing required column: '{col_name}'. "
                f"Found: {', '.join(col_idx.keys())}"
            )
            wb.close()
            return result

    if not batch_label:
        batch_label = f"excel-{timezone.now().strftime('%Y%m%d-%H%M%S')}"

    # Optionally withdraw existing available properties
    if replace_existing:
        Property.objects.filter(status=Property.Status.AVAILABLE).update(
            status=Property.Status.WITHDRAWN
        )

    idx_addr = col_idx["Street Address"]
    idx_city = col_idx.get("City State Zip")
    idx_owned = col_idx.get("GCLB Owned")
    idx_structure = col_idx.get("Structure Flag")
    idx_bid = col_idx.get("Minimum Bid")
    idx_program = col_idx.get("Program Type")

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue

        # Filter: only GCLB-owned properties
        owned = str(row[idx_owned]).strip() if idx_owned is not None and row[idx_owned] else ""
        if owned != "Yes":
            result["skipped"] += 1
            continue

        # Extract address
        street = str(row[idx_addr]).strip() if row[idx_addr] else ""
        if not street:
            result["errors"].append(f"Row {row_num}: missing Street Address")
            continue

        # Combine street + city for a complete address
        city = ""
        if idx_city is not None and row[idx_city]:
            city = str(row[idx_city]).strip()
        address = f"{street}, {city}" if city else street

        # Program type
        program_raw = str(row[idx_program]).strip() if idx_program is not None and row[idx_program] else ""
        structure = str(row[idx_structure]).strip() if idx_structure is not None and row[idx_structure] else ""
        program_type = _infer_program_type(program_raw, structure)
        if not program_type:
            result["errors"].append(
                f"Row {row_num}: cannot determine program type "
                f"(program='{program_raw}', structure='{structure}')"
            )
            continue

        # Listing price from Minimum Bid
        listing_price = None
        if idx_bid is not None and row[idx_bid]:
            listing_price = _parse_price(str(row[idx_bid]))
            # Skip zero bids — they're not real prices
            if listing_price is not None and listing_price == 0:
                listing_price = None

        # Upsert by address_normalized (no parcel_id in FileMaker export)
        normalized = Property.normalize_address(address)
        try:
            prop, created = Property.objects.update_or_create(
                address_normalized=normalized,
                defaults={
                    "address": address,
                    "program_type": program_type,
                    "listing_price": listing_price,
                    "status": Property.Status.AVAILABLE,
                    "csv_batch": batch_label,
                },
            )
            if created:
                result["created"] += 1
            else:
                result["updated"] += 1
        except Exception as e:
            result["errors"].append(f"Row {row_num}: {e}")

    wb.close()
    return result
